"""End-to-end coverage for the three imports exposed by the core list views."""

import io

from openpyxl import load_workbook


def _template_with_row(template_bytes: bytes, values: dict) -> bytes:
    workbook = load_workbook(io.BytesIO(template_bytes))
    sheet = workbook["Data"]
    columns = {cell.value: cell.column for cell in sheet[1]}
    for name, value in values.items():
        sheet.cell(row=2, column=columns[name], value=value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_function_task_and_gantt_templates_can_be_imported(auth_client):
    project = auth_client.post(
        "/api/projects",
        json={"name": "Core import API", "project_code": "CIA"},
    )
    assert project.status_code == 200, project.text
    slug = project.json()["slug"]

    cases = (
        ("functions", {"name": "Import function", "type": "Functional", "phase": "UR"}),
        ("tasks", {"title": "Import task", "phase": "UR", "is_followup": False}),
        (
            "gantt",
            {"name": "Import schedule", "phase": "UR", "start_date": "2026-08-31", "end_date": "2026-09-04"},
        ),
    )

    for entity, values in cases:
        template = auth_client.get(f"/api/{slug}/{entity}/import-template")
        assert template.status_code == 200, template.text
        upload = _template_with_row(template.content, values)
        response = auth_client.post(
            f"/api/{slug}/{entity}/import",
            files={"file": (f"{entity}.xlsx", upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200, response.text
        assert response.json()["imported"] == 1


def test_progress_matrix_schedule_template_imports_and_exports_function_dates(auth_client):
    project = auth_client.post(
        "/api/projects",
        json={"name": "Matrix schedule import", "project_code": "MSI"},
    )
    assert project.status_code == 200, project.text
    slug = project.json()["slug"]

    function = auth_client.post(
        f"/api/{slug}/functions",
        json={"function_code": "MSI-FN-001", "name": "Dated function"},
    )
    assert function.status_code == 200, function.text

    template = auth_client.get(f"/api/{slug}/schedule/import-template")
    assert template.status_code == 200, template.text
    upload = _template_with_row(
        template.content,
        {
            "entity_type": "function",
            "entity_code": "MSI-FN-001",
            "plan_start": "2026-09-01",
            "plan_end": "2026-09-10",
            "actual_start_override": "2026-09-02",
            "actual_end_override": "2026-09-11",
            "override_reason": "Migrated from the prior plan",
        },
    )
    imported = auth_client.post(
        f"/api/{slug}/schedule/import",
        files={"file": ("schedule.xlsx", upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["plan_dates_set"] == 1
    assert imported.json()["actual_overrides_set"] == 1

    matrix = auth_client.get(f"/api/{slug}/progress-matrix", params={"entity_type": "function"})
    assert matrix.status_code == 200, matrix.text
    row = next(row for row in matrix.json()["rows"] if row["entity_code"] == "MSI-FN-001")
    assert row["plan_start"] == "2026-09-01"
    assert row["plan_end"] == "2026-09-10"
    assert row["actual_start_override"] == "2026-09-02"
    assert row["actual_end_override"] == "2026-09-11"

    exported = auth_client.get(f"/api/{slug}/schedule/export")
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    rows = list(workbook["Data"].iter_rows(values_only=True))
    assert rows[0] == (
        "entity_type",
        "entity_code",
        "plan_start",
        "plan_end",
        "actual_start_override",
        "actual_end_override",
        "override_reason",
        "actual_start_derived",
        "actual_end_derived",
        "actual_start",
        "actual_end",
        "actual_start_source",
        "actual_end_source",
    )
    function_row = next(row for row in rows[1:] if row[0] == "function" and row[1] == "MSI-FN-001")
    assert function_row[7:13] == (None, None, "2026-09-02", "2026-09-11", "override", "override")
